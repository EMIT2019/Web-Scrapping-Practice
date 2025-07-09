from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

class ExcelReader:
    def __init__(self, file_path:str) -> None:
        self.file_path = file_path
        self.file_instance:Workbook = None
        self.worksheet:Worksheet = None
        self.sheetnames:list[str] = []
        self.articles_list:list[str] = []

    def load_file_instance(self) -> None:
        try:
            self.file_instance = load_workbook(self.file_path)
        except Exception as ex:
            print(f"Something went wrong during the execution. This is the full error message: {ex}")

    def get_sheetnames(self) -> None:
        try:
            self.sheetnames = self.file_instance.sheetnames
        except Exception as ex:
            print(f"Something went wrong during the execution. This is the full error message: {ex}")

    def load_worksheet(self) -> None:
        try:
            self.worksheet = self.file_instance[self.sheetnames[0]]
        except Exception as ex:
            print(f"Something went wrong during the execution. This is the full error message: {ex}")

    def set_articles_list(self) -> None:
        try:
            for row in self.worksheet.iter_rows():
                for cell in row:
                    if cell.col_idx == 1 and cell.row > 1:
                        self.articles_list.append(cell.value)

        except Exception as ex:
            print(f"Something went wrong during the execution. This is the full error message: {ex}")
